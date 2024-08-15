# Excel2IDS
# Copyright (C) 2024 Artur Tomczak <artomczak@gmail.com>

import os
import sys
import time
import copy
import openpyxl
from ifctester import ids
from tqdm import tqdm
import json
import re
import datetime


class Settings:
    def __init__(self, settings_file):
        with open(settings_file, 'r') as file:
            settings_dict = json.load(file)
        for key, value in settings_dict.items():
            setattr(self, key, value)


def isempty(v):
    return (v is None or v == '')


def excel2ids(spreadsheet, ids_path):
    sheet = spreadsheet[s.SHEET_NAME]
    # Define the starting cell of the assignment table
    start_row = sheet[s.DEFAULT_START_CELL].row
    start_col = sheet[s.DEFAULT_START_CELL].column
    end_row = sheet.max_row + 1
    end_col = sheet.max_column + 1

    for col in tqdm(range(start_col, end_col), desc="Processing Excel columns."):
        column_letter = openpyxl.utils.get_column_letter(col)

        if sheet[f'{column_letter}{s.APL_INCLUDE}'].value:
            ### add applicability
            applicability = []
            # add entity and predefined type
            entity_cell = process_value(sheet[f'{column_letter}{s.APL_ENTITY}'].value)
            predefined_type_cell = process_value(sheet[f'{column_letter}{s.APL_PRED_TYPE}'].value)
            if entity_cell:
                if predefined_type_cell:
                    entity = ids.Entity(name=entity_cell, predefinedType=predefined_type_cell)
                else:
                    entity = ids.Entity(name=entity_cell)
                # if '.' in entity_cell:
                #     entity = ids.Entity(name=entity_cell.split('.')[0].upper(), predefinedType=entity_cell.split('.')[1].upper())
                # else:
                applicability.append(entity)
            # add property
            if sheet[f'{column_letter}{s.APL_PNAME}'].value:
                property = ids.Property(
                    propertySet=process_value(sheet[f'{column_letter}{s.APL_PSET}'].value), 
                    baseName=process_value(sheet[f'{column_letter}{s.APL_PNAME}'].value),
                    dataType=process_value(sheet[f'{column_letter}{s.APL_PDTYPE}'].value)
                    )             
                pv = sheet[f"{column_letter}{s.APL_PVAL}"].value
                if not isempty(pv):
                    property.value = process_value(pv)
                applicability.append(property)
            # add classification
            if not isempty(sheet[f'{column_letter}{s.APL_CLASS_SYS}'].value):
                classification = ids.Classification(
                    system=process_value(sheet[f'{column_letter}{s.APL_CLASS_SYS}'].value),
                    value=process_value(sheet[f'{column_letter}{s.APL_CLASS_CODE}'].value))
                applicability.append(classification)
            # add attribute
            if not isempty(sheet[f'{column_letter}{s.APL_ANAME}'].value):
                attribute = ids.Attribute(
                    name=process_value(sheet[f'{column_letter}{s.APL_ANAME}'].value),
                    value=process_value(sheet[f'{column_letter}{s.APL_AVALUE}'].value)
                )
                applicability.append(attribute)
            # add material
            if not isempty(sheet[f'{column_letter}{s.APL_MATERIAL}'].value):
                material = ids.Material(
                    value=process_value(sheet[f'{column_letter}{s.APL_MATERIAL}'].value)
                )
                applicability.append(material)

            ### add requirement(s)
            requirements = []
            for row in range(start_row, end_row):
                if sheet[f'{s.REQ_INCLUDE}{row}'].value:
                    cell_value = sheet.cell(row=row, column=col).value
                    if not isempty(cell_value):
                        if cell_value.strip().upper() == "X":
                            instructions = sheet[f'{s.REQ_INSTRUCTIONS}{row}'].value
                            # add entity and predefined type
                            entity_cell = process_value(sheet[f'{s.REQ_ENTITY}{row}'].value)
                            predefined_type_cell = process_value(sheet[f'{s.REQ_PRED_TYPE}{row}'].value)
                            if entity_cell:
                                if predefined_type_cell:
                                    entity = ids.Entity(name=entity_cell, predefinedType=predefined_type_cell)
                                else:
                                    entity = ids.Entity(name=entity_cell)
                                if instructions:
                                    entity.instructions=instructions
                                requirements.append(entity)
                            # add property
                            if not isempty(sheet[f'{s.REQ_PNAME}{row}'].value):
                                property = ids.Property(
                                    propertySet=process_value(sheet[f'{s.REQ_PSET}{row}'].value), 
                                    baseName=process_value(sheet[f'{s.REQ_PNAME}{row}'].value),
                                    dataType=process_value(sheet[f'{s.REQ_PDTYPE}{row}'].value),
                                    cardinality=process_value(sheet[f'{s.REQ_CARDINAL}{row}'].value)
                                )
                                if not isempty(sheet[f'{s.REQ_PVAL}{row}'].value):
                                    property.value = process_value(sheet[f'{s.REQ_PVAL}{row}'].value)
                                if not isempty(sheet[f'{s.REQ_URI}{row}'].value):
                                    property.uri = process_value(sheet[f'{s.REQ_URI}{row}'].value)
                                if instructions:
                                    property.instructions=instructions
                                requirements.append(property)
                            # add classification
                            if not isempty(sheet[f'{s.REQ_CLASS_SYS}{row}'].value):
                                classification = ids.Classification(
                                    system=process_value(sheet[f'{s.REQ_CLASS_SYS}{row}'].value),
                                    value=process_value(sheet[f'{s.REQ_CLASS_CODE}{row}'].value),
                                    cardinality=process_value(sheet[f'{s.REQ_CARDINAL}{row}'].value)
                                )
                                if not isempty(sheet[f'{s.REQ_CLASS_URI}{row}'].value):
                                    classification.uri = process_value(sheet[f'{s.REQ_CLASS_URI}{row}'].value)
                                if instructions:
                                    classification.instructions=instructions
                                requirements.append(classification)
                            # add attribute
                            if not isempty(sheet[f'{s.REQ_ANAME}{row}'].value):
                                attribute = ids.Attribute(
                                    name=process_value(sheet[f'{s.REQ_ANAME}{row}'].value),
                                    value=process_value(sheet[f'{s.REQ_AVALUE}{row}'].value),
                                    cardinality=process_value(sheet[f'{s.REQ_CARDINAL}{row}'].value)
                                )
                                if instructions:
                                    attribute.instructions=instructions
                                requirements.append(attribute)
                            # add material
                            if not isempty(sheet[f'{s.REQ_MATERIAL}{row}'].value):
                                material = ids.Material(
                                    value=process_value(sheet[f'{s.REQ_MATERIAL}{row}'].value),
                                    cardinality=process_value(sheet[f'{s.REQ_CARDINAL}{row}'].value)
                                )
                                if instructions:
                                    material.instructions=instructions
                                requirements.append(material)         
                        else:
                            # process 'REPLACEME'
                            cell_value = process_value(cell_value)
                            if sheet[f'{s.REQ_ENTITY}{row}'].value == 'REPLACEME':
                                if predefined_type_cell:
                                    facet = ids.Entity(name=cell_value, predefinedType=predefined_type_cell)
                                else:
                                    facet = ids.Entity(name=cell_value)
                            elif sheet[f'{s.REQ_PSET}{row}'].value == 'REPLACEME':
                                facet = ids.Property(
                                    propertySet=cell_value, 
                                    baseName=process_value(sheet[f'{s.REQ_PNAME}{row}'].value),
                                    dataType=process_value(sheet[f'{s.REQ_PDTYPE}{row}'].value),
                                    cardinality=process_value(sheet[f'{s.REQ_CARDINAL}{row}'].value)
                                )
                                if not isempty(sheet[f'{s.REQ_PVAL}{row}'].value):
                                    property.value = process_value(sheet[f'{s.REQ_PVAL}{row}'].value)
                                if not isempty(sheet[f'{s.REQ_URI}{row}'].value):
                                    property.uri = process_value(sheet[f'{s.REQ_URI}{row}'].value)
                            elif sheet[f'{s.REQ_PNAME}{row}'].value == 'REPLACEME':
                                facet = ids.Property(
                                    propertySet=process_value(sheet[f'{s.REQ_PSET}{row}'].value), 
                                    baseName=cell_value,
                                    dataType=process_value(sheet[f'{s.REQ_PDTYPE}{row}'].value),
                                    cardinality=process_value(sheet[f'{s.REQ_CARDINAL}{row}'].value)
                                )
                                if not isempty(sheet[f'{s.REQ_PVAL}{row}'].value):
                                    property.value = process_value(sheet[f'{s.REQ_PVAL}{row}'].value)
                                if not isempty(sheet[f'{s.REQ_URI}{row}'].value):
                                    property.uri = process_value(sheet[f'{s.REQ_URI}{row}'].value)
                            elif sheet[f'{s.REQ_PVAL}{row}'].value == 'REPLACEME':
                                facet = ids.Property(
                                    propertySet=process_value(sheet[f'{s.REQ_PSET}{row}'].value), 
                                    baseName=process_value(sheet[f'{s.REQ_PNAME}{row}'].value),
                                    dataType=process_value(sheet[f'{s.REQ_PDTYPE}{row}'].value),
                                    cardinality=process_value(sheet[f'{s.REQ_CARDINAL}{row}'].value),
                                    value = cell_value
                                )
                                if not isempty(sheet[f'{s.REQ_URI}{row}'].value):
                                    property.uri = process_value(sheet[f'{s.REQ_URI}{row}'].value)
                            elif sheet[f'{s.REQ_CLASS_SYS}{row}'].value == 'REPLACEME':
                                facet = ids.Classification(
                                    system=cell_value,
                                    value=process_value(sheet[f'{s.REQ_CLASS_CODE}{row}'].value),
                                    cardinality=process_value(sheet[f'{s.REQ_CARDINAL}{row}'].value)
                                    )
                                if not isempty(sheet[f'{s.REQ_CLASS_URI}{row}'].value):
                                    facet.uri = process_value(sheet[f'{s.REQ_CLASS_URI}{row}'].value)
                            elif sheet[f'{s.REQ_CLASS_CODE}{row}'].value == 'REPLACEME':
                                facet = ids.Classification(
                                    system=process_value(sheet[f'{s.REQ_CLASS_SYS}{row}'].value),
                                    value=cell_value,
                                    cardinality=process_value(sheet[f'{s.REQ_CARDINAL}{row}'].value)
                                    )
                                if not isempty(sheet[f'{s.REQ_CLASS_URI}{row}'].value):
                                    facet.uri = process_value(sheet[f'{s.REQ_CLASS_URI}{row}'].value)
                            elif sheet[f'{s.REQ_ANAME}{row}'].value == 'REPLACEME':
                                facet = ids.Attribute(
                                    name=cell_value,
                                    value=process_value(sheet[f'{s.REQ_AVALUE}{row}'].value),
                                    cardinality=process_value(sheet[f'{s.REQ_CARDINAL}{row}'].value)
                                )       
                            elif sheet[f'{s.REQ_AVALUE}{row}'].value == 'REPLACEME':
                                facet = ids.Attribute(
                                    name=process_value(sheet[f'{s.REQ_ANAME}{row}'].value),
                                    value=cell_value,
                                    cardinality=process_value(sheet[f'{s.REQ_CARDINAL}{row}'].value)
                                )                                          
                            elif sheet[f'{s.REQ_MATERIAL}{row}'].value == 'REPLACEME':
                                facet = ids.Material(value=cell_value, cardinality=process_value(sheet[f'{s.REQ_CARDINAL}{row}'].value))
                            
                            if facet:
                                if instructions:
                                    facet.instructions=instructions
                                requirements.append(facet)
                            else:
                                print(color_text(f"The only allowed values are 'X' and 'REPLACEME' but your table has: '{cell_value}'.", color='red'))

            if requirements:
                disciplines = split_multivalue(sheet[f"{column_letter}{s.APL_PURPOSE}"].value)
                ifc_version = sheet[s.IFC_VERSION].value
                if not ifc_version in ['IFC2X3','IFC4','IFC4X3_ADD2']:
                    ifc_version = s.IFC_VERSION_DEFAULT
                for discipline in disciplines:
                    add_to_ids(
                        applicability,
                        requirements,
                        apl_cardinality=sheet[f'{column_letter}{s.APL_CARDINAL}'].value,
                        purpose=discipline,
                        spec_name=sheet[f"{column_letter}{s.SPE_NAME}"].value,
                        spec_description=sheet[f"{column_letter}{s.SPE_DESCR}"].value,
                        spec_instructions=sheet[f"{column_letter}{s.SPE_INSTR}"].value,
                        spec_identifier=sheet[f"{column_letter}{s.SPE_IDENT}"].value,
                        #TODO process milestone/phases REQ_MILESTONE
                        milestone=MILESTONE,
                        ifc_version=ifc_version,
                        title=sheet[s.IDS_TITLE].value,
                        author=sheet[s.IDS_AUTHOR].value,
                        date=sheet[s.IDS_DATE].value,
                        version=sheet[s.IDS_VERSION].value,
                        copyright=sheet[s.IDS_COPYRIGHT].value,
                        description=sheet[s.IDS_DESCRIPTION].value,
                    )

    ### Save all IDSes to files:
    for new_ids in tqdm(ids_list, desc="Generating separate .ids files."):
        ids_list[new_ids].to_xml(ids_path.replace(".ids", "_" + new_ids + ".ids"))
    print(
        color_text(f"Success! {len(ids_list)} IDS files were saved in {os.path.dirname(ids_path)}.", color='green')
    )

    # Close the spreadsheet
    spreadsheet.close()


def add_to_ids(
    applicability,
    requirements,
    apl_cardinality="required",
    purpose="General",
    milestone="Handover",
    spec_name="Specification name",
    spec_description="",
    spec_instructions="",
    spec_identifier="",
    ifc_version="IFC4X3_ADD2",
    title="Generic IDS title",
    author="Anonymous",
    date=datetime.date.today(),
    version="0.1",
    copyright="No copyright",
    description="",
):
    """Add this specifiction to IDS file. If such IDS doesn't exist yet, create it.
    TODO Known limitations: 
    - the applicability is automatically set to 'minOccur'=0, meaning 'if exists'/'may occur' and does not trigger an error if no such element is found.
    """

    if isinstance(date,datetime.datetime):
        date = date.strftime("%Y-%m-%d")

    if not purpose in ids_list:
        # create new IDS
        ids_list[purpose] = ids.Ids(
            title=title,
            author=author,
            version=str(version),
            description=description,
            copyright=copyright,
            date=date,
            purpose=purpose,
            milestone=milestone
        )

    # check if this IDS already has such category (applicability)
    exists = False
    for spec in ids_list[purpose].specifications:
        if spec.name == spec_name:
            # ADD req!
            spec.requirements += requirements
            exists = True

    if not exists:
        # create new spec
        new_spec = ids.Specification(
            name=spec_name,
            minOccurs=1,
            maxOccurs='unbounded',
            ifcVersion=ifc_version,
            identifier=spec_identifier,
            description=spec_description,
            instructions=spec_instructions,
        )

        if apl_cardinality == 'prohibited':
            new_spec.minOccurs=0
            new_spec.maxOccurs=0
        elif apl_cardinality == 'optional':
            new_spec.minOccurs=0
            new_spec.maxOccurs='unbounded'

        new_spec.applicability = copy.deepcopy(applicability)
        new_spec.requirements = copy.deepcopy(requirements)
        ids_list[purpose].specifications.append(new_spec)


QUOTED_PATTERN = r'^".*"$'

def process_value(cell_value):
    """ Look at the value and convert to enumeration or pattern if needed """
    if not isempty(cell_value):
        if isinstance(cell_value, str):
            # remove trailing spaces
            cell_value=cell_value.strip()
            # if it's in quotation, turn into pattern restriction (regex):
            if bool(re.match(QUOTED_PATTERN, cell_value)):
                cell_value = ids.Restriction(options={"pattern": cell_value[1:-1]})
            # if there are multiple lines in a single cell, split it into enumeration of literal values
            elif "\n" in cell_value or "," in cell_value or ";" in cell_value:
                cell_value = ids.Restriction(options={"enumeration": split_multivalue(cell_value)})
            # in other cases, return as is (empty or literal value)
        elif isinstance(cell_value, bool):
            cell_value=str(cell_value)
    return cell_value


SPLIT_PATTERN=r'\s*[,;\n]\s*'

def split_multivalue(text):
    """ Split discipline/phase text if plural (allowed multiline, comma-separated and semicolon-separated).  """
    return re.split(SPLIT_PATTERN, text)


def color_text(text, color='blue'):
    if color == 'blue':
        text = '\033[94m' + text + '\033[0m'
    elif color == 'red':
        text = '\033[31m' + text + '\033[0m'
    elif color == 'green':
        text = '\033[92m' + text + '\033[0m'
    return text


def ask_for_path():
    file_path = input(color_text("\nPlease enter the path to the Excel spreadsheet: \n"))
    if file_path[0] == '"':
        file_path = file_path[1:]
    if file_path[-1] == '"':
        file_path = file_path[:-1]
    if file_path[-5:] != '.xlsx':
        print(color_text("\nThe file must be an .xlsx. Please check the path and try again.", color='red'))
        ask_for_path()
    try:
        spreadsheet = openpyxl.load_workbook(file_path)
        return spreadsheet, file_path
    except FileNotFoundError:
        print(color_text("\nThe file was not found. Please check the path and try again.", color='red'))
        ask_for_path()
    except Exception as e:
        print(color_text(f"\nAn error occurred: {e}", color='red'))
        print(color_text("\nThe program will close automatically in 10 seconds...\n"))
        time.sleep(10)
        sys.exit()


if __name__ == "__main__": 

    s = Settings('settings.json')
    ids_list = {}

    spreadsheet, file_path = ask_for_path()
    ids_path = file_path.replace(".xlsx", ".ids")

    MILESTONE = 'LOD400' #TODO TEMP workaround for phases/milestones

    excel2ids(spreadsheet, ids_path)

    time.sleep(1)
    print(color_text("\nThe program will close automatically in 5 seconds...\n"))
    time.sleep(6)
    sys.exit()